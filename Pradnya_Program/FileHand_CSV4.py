print("Determine total number of rows and columns")
import openpyxl
file_name = r"D:\\python class\\Programs\\Pradnya_Program\\Openpyxl Data1.xlsx"

wb = openpyxl.load_workbook(file_name)

ws = wb.active

print(f"total rows are {ws.max_row}")

print(f"total columns are {ws.max_column}")
#-------------------------------------------------------------------
print("To print all column names")
col_max = ws.max_column

for i in range(1,(col_max +1)):
    cell_read = ws.cell(row = 1, column = i)
    print(cell_read.value)
