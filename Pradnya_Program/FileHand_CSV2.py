from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws1 = wb.create_sheet("sheet1",0)
ws2 = wb.create_sheet("example2",2)
sw3 = wb.create_sheet("example3",-1)

ws.title = "example1"
ws1.title = "example0"
wb.save(filename = "example.xlsx")
